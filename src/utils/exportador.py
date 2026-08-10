"""
exportador.py — Genera Excel del cashflow completo
Con formato profesional, fórmulas y estilos.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os, sys
sys.path.insert(0, '.')
from config import EMPRESA, AÑO, EXPORT_DIR
from src.utils.helpers import fmt_ars, logger

# Paleta
C_DARK   = "1F3864"; C_MED = "2E75B6"; C_LIGHT = "BDD7EE"
C_GREEN  = "E2EFDA"; C_RED = "FFDDE1"; C_TOTAL = "FFE699"
C_INPUT  = "FFFDE7"; C_FORM = "F2F2F2"; C_AGUINALDO = "FFF2CC"

def fill(h): return PatternFill("solid", fgColor=h)
def borde():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, r, c, val=None, bold=False, bg=None, fc="000000", ah="left", num=None, sz=10):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=bold, color=fc, size=sz, name="Arial")
    if bg: cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal=ah, vertical="center")
    cell.border = borde()
    if num: cell.number_format = num
    return cell

FMT_PESOS = '#.##0;(#.##0);"-"'
MESES_CORTO = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

def exportar_excel(
    df_cashflow: pd.DataFrame,
    df_extracto: pd.DataFrame = None,
    path_salida: str = None
) -> str:
    """Genera Excel completo del cashflow."""
    if path_salida is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        path_salida = os.path.join(
            "./exports",
            f"cashflow_{AÑO}_{ts}.xlsx"
        )
    os.makedirs(os.path.dirname(path_salida), exist_ok=True)
    wb = Workbook()

    # ── Hoja 1: CASHFLOW MENSUAL ──────────────────────────────────────
    ws = wb.active
    ws.title = "CASHFLOW_MENSUAL"
    ws.column_dimensions["A"].width = 26
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Título
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = f"CASHFLOW MENSUAL {AÑO} — {EMPRESA}"
    c.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    c.fill = fill(C_DARK); c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # Generado
    ws.merge_cells("A2:N2")
    ws["A2"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Sistema Cashflow Inteligente v1.0"
    ws["A2"].font = Font(italic=True, color="555555", size=9, name="Arial")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Header meses
    headers = ["Concepto"] + MESES_CORTO + ["TOTAL"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        c.fill = fill(C_DARK); c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borde()

    # Datos
    filas = [
        # (label, col_df, bg, bold)
        ("── SALDO INICIAL", None, "D6E4F7", True),
        ("Saldo Inicial del Mes",       "saldo_ini_proy", C_TOTAL, True),
        ("── INGRESOS",                 None, "D6E4F7", True),
        ("Cobros Proyectados",          "ing_proy",    C_FORM, False),
        ("  ▸ Contado",                 "cobro_contado", C_FORM, False),
        ("  ▸ 30 días",                 "cobro_30d",   C_FORM, False),
        ("  ▸ 60 días",                 "cobro_60d",   C_FORM, False),
        ("  ▸ 90 días",                 "cobro_90d",   C_FORM, False),
        ("Cobros Reales",               "ing_real",    C_INPUT, False),
        ("Desvío Ingresos $",           "dev_ing",     C_FORM, False),
        ("── EGRESOS",                  None, "D6E4F7", True),
        ("Sueldos Brutos",              "sueldos",     C_FORM, False),
        ("Cargas Sociales",             "cargas_sociales", C_FORM, False),
        ("IVA / AFIP",                  "iva",         C_FORM, False),
        ("Cuotas Préstamos",            "cuotas_prestamos", C_FORM, False),
        ("Proveedores",                 "proveedores", C_FORM, False),
        ("── RESULTADO",                None, "D6E4F7", True),
        ("Resultado Neto Proyectado",   "res_proy",    C_TOTAL, True),
        ("Resultado Neto Real",         "res_real",    C_INPUT, True),
        ("── SALDO FINAL",              None, "D6E4F7", True),
        ("Saldo Final Proyectado",      "saldo_fin_proy", C_MED, True),
        ("Saldo Final Real",            "saldo_fin_real", C_INPUT, True),
        ("Semáforo",                    "semaforo",    C_FORM, False),
    ]

    totales_excluir = {"semaforo", None}
    row_num = 4
    for label, col_df, bg, bold in filas:
        if col_df is None:
            # Sección
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=14)
            c = ws.cell(row=row_num, column=1, value=label)
            c.font = Font(bold=True, color=C_DARK, size=10, name="Arial")
            c.fill = fill(bg); c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = borde()
            row_num += 1
            continue

        sc(ws, row_num, 1, label, bold=bold, bg=bg, ah="left")
        total_row = 0
        for mi, mes_n in enumerate(range(1, 13)):
            col = 2 + mi
            row_data = df_cashflow[df_cashflow["mes"] == mes_n]
            if not row_data.empty and col_df in row_data.columns:
                val = row_data.iloc[0][col_df]
                # Aguinaldo highlight
                bg_cell = C_AGUINALDO if (col_df == "cargas_sociales" and mes_n in [6,12]) else bg
                c = sc(ws, row_num, col, val if val is not None else None,
                       bold=bold, bg=bg_cell, ah="center", num=FMT_PESOS if col_df != "semaforo" else None)
                if val and col_df not in totales_excluir and isinstance(val, (int, float)):
                    total_row += float(val)
            else:
                sc(ws, row_num, col, None, bg=bg, ah="center")

        # Total anual (col 14)
        if col_df not in totales_excluir and col_df not in ["saldo_ini_proy","saldo_fin_proy","saldo_fin_real","semaforo"]:
            sc(ws, row_num, 14, total_row, bold=True, bg=C_TOTAL, ah="center", num=FMT_PESOS)
        else:
            sc(ws, row_num, 14, None, bg=C_FORM, ah="center")

        row_num += 1

    # ── Hoja 2: EXTRACTO (si existe) ─────────────────────────────────
    if df_extracto is not None and not df_extracto.empty:
        ws2 = wb.create_sheet("EXTRACTO_BANCARIO")
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 14
        ws2.column_dimensions["C"].width = 42
        ws2.column_dimensions["D"].width = 16
        ws2.column_dimensions["E"].width = 18
        ws2.column_dimensions["F"].width = 20
        ws2.column_dimensions["G"].width = 16

        ws2.merge_cells("A1:G1")
        c = ws2["A1"]
        c.value = f"EXTRACTO BANCARIO — {df_extracto['banco'].iloc[0] if len(df_extracto) > 0 else ''}"
        c.font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
        c.fill = fill(C_DARK); c.alignment = Alignment(horizontal="center")

        hdrs = ["N°","Fecha","Descripción","Importe $","Categoría","Banco","Estado Conc."]
        for j, h in enumerate(hdrs, 1):
            c = ws2.cell(row=2, column=j, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
            c.fill = fill(C_MED); c.alignment = Alignment(horizontal="center")
            c.border = borde()

        for idx, (_, row) in enumerate(df_extracto.iterrows(), 3):
            bg_row = C_GREEN if row.get("importe", 0) > 0 else C_RED
            sc(ws2, idx, 1, idx-2, bg=C_FORM, ah="center")
            sc(ws2, idx, 2, row.get("fecha_str",""), bg=C_FORM, ah="center")
            sc(ws2, idx, 3, row.get("descripcion",""), bg="FFFFFF", ah="left")
            sc(ws2, idx, 4, row.get("importe"), bg=bg_row, ah="center", num=FMT_PESOS)
            sc(ws2, idx, 5, row.get("categoria",""), bg=C_FORM, ah="center")
            sc(ws2, idx, 6, row.get("banco",""), bg=C_FORM, ah="center")
            estado = row.get("estado_conciliacion", "")
            sc(ws2, idx, 7, estado, bg=C_FORM, ah="center")

    wb.save(path_salida)
    logger.ok(f"Excel exportado: {path_salida}")
    return path_salida


if __name__ == "__main__":
    from src.engine.motor_cashflow import (
        ParametrosCashflow, generar_cashflow_mensual
    )
    from src.parsers.parser_bancario import generar_extracto_muestra, parse_extracto

    params = ParametrosCashflow()
    df_m = generar_extracto_muestra()
    path_m = "./data/samples/test.csv"
    df_m.to_csv(path_m, index=False)
    df_real = parse_extracto(path_m, banco="nacion")
    df_cf = generar_cashflow_mensual(params, df_real=df_real)

    path_out = exportar_excel(df_cf, df_real)
    print(f"✅ Excel generado: {path_out}")
